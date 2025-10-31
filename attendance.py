import cv2
import pytesseract
import numpy as np
import re
from tkinter import *
from tkinter import filedialog
from openpyxl import Workbook
from PIL import Image, ImageTk

pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

def save_to_excel(registration_numbers, excel_file="Scanned_Registration_Numbers.xlsx"):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Registration Numbers"
    sheet.cell(row=1, column=1).value = "Registration Number"

    for idx, reg_num in enumerate(registration_numbers, start=2):
        sheet.cell(row=idx, column=1).value = reg_num

    workbook.save(excel_file)
    print(f"Saved {len(registration_numbers)} registration numbers to {excel_file}")

def preprocess_image(frame):
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    blurred = cv2.GaussianBlur(gray, (5, 5), 0)
    thresh = cv2.adaptiveThreshold(
        blurred, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY_INV, 11, 2
    )

    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (3, 3))
    dilated = cv2.dilate(thresh, kernel, iterations=1)
    return dilated

def detect_text_regions(preprocessed_frame):
    contours, _ = cv2.findContours(preprocessed_frame, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    regions = []

    for contour in contours:
        x, y, w, h = cv2.boundingRect(contour)
        aspect_ratio = w / float(h)

        if 100 < w < 1000 and 20 < h < 200 and 2.0 < aspect_ratio < 6.0:
            regions.append((x, y, w, h))

    return regions

def perspective_correction(frame, region):
    x, y, w, h = region
    margin = 10
    roi = frame[max(0, y - margin):y + h + margin, max(0, x - margin):x + w + margin]

    if roi.size == 0:
        return None

    return roi

def extract_text_from_region(region):
    text = pytesseract.image_to_string(region, config='--psm 6')  # Page Segmentation Mode 6: Assume uniform text block
    return text

def extract_registration_numbers(text):
    matches = re.findall(r"\b\d{2}[A-Z]{3}\d{5}\b", text)  # Example: 22BCY10111
    return matches

class RegistrationScannerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Registration Number Scanner")

        self.registration_numbers = set()

        self.video_label = Label(root)
        self.video_label.pack()

        self.start_button = Button(root, text="Start Scan", command=self.start_scan)
        self.start_button.pack()

        self.save_button = Button(root, text="Save to Excel", command=self.save_to_excel)
        self.save_button.pack()

        self.quit_button = Button(root, text="Quit", command=root.quit)
        self.quit_button.pack()

        self.cap = None

    def start_scan(self):
        self.cap = cv2.VideoCapture(0)  
        self.process_frame()

    def process_frame(self):
        ret, frame = self.cap.read()
        if not ret:
            print("Failed to capture frame")
            return

        preprocessed_frame = preprocess_image(frame)

        text_regions = detect_text_regions(preprocessed_frame)

        for region in text_regions:
            x, y, w, h = region
            cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)

            corrected_region = perspective_correction(preprocessed_frame, region)
            if corrected_region is not None:
                text = extract_text_from_region(corrected_region)
                registration_numbers = extract_registration_numbers(text)

                if registration_numbers:
                    self.registration_numbers.update(registration_numbers)
                    print(f"Detected: {registration_numbers}")

        for region in text_regions:
            x, y, w, h = region
            cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)

        cv2_im = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        img = Image.fromarray(cv2_im)
        imgtk = ImageTk.PhotoImage(image=img)

        self.video_label.imgtk = imgtk
        self.video_label.configure(image=imgtk)

        self.video_label.after(10, self.process_frame)

    def save_to_excel(self):
        if self.registration_numbers:
            file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
            if file_path:
                save_to_excel(self.registration_numbers, file_path)
            self.registration_numbers.clear()  
            print("No registration numbers detected.")

if __name__ == "__main__":
    root = Tk()
    app = RegistrationScannerApp(root)
    root.mainloop()